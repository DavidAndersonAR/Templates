import openpyxl
import pymysql


# Abrir o arquivo Excel
arquivo_excel = "orion-abril.xlsx"
planilha = openpyxl.load_workbook(arquivo_excel).active


def conecta_no_banco():
    host = "52.177.123.37"
    user = "root"
    password = "enelsafe2016"
    database = "zabbix_enel"
    port = 3306

    conexao = pymysql.connect(host=host, user=user, password=password, database=database, port=port)
    cursor = conexao.cursor()

    return cursor

def adiciona_linha():
    cursor = conecta_no_banco()

    contador = 0
    for linha in planilha.iter_rows(min_row=2, values_only=True):
        id, localidade, host, *dados = linha
        # Iterar sobre os dados
        for indice, valor in enumerate(dados):
            data_coluna = planilha.cell(row=1, column=indice + 4).value  # A coluna começa da quarta coluna
            contador = contador +1
            data_coluna = data_coluna.date()
        
            # if valor is not None:
            #     porcentagem_formatada = "{:.2f}".format(valor * 100).replace('.', ',')
                
            # else:
            #     porcentagem_formatada = 0,00

            sql = "INSERT INTO disponibilidade_hosts (id, host, data_coleta, disp_expurgada, status) VALUES (%s, %s, %s, %s, %s)"
            valores = (id, host, data_coluna, valor, "Ativo")

            try:
                cursor.execute(sql, valores)
                #print(valores)
                print(f"{contador} -- Comando SQL executado com sucesso {sql} - {host}")
            except Exception as erro:
                print(erro)
        cursor.connection.commit()
    cursor.connection.close()

adiciona_linha()